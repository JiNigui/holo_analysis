#!/usr/bin/env python3
"""
导线熔痕三维孔洞全参数特征提取与智能鉴定引擎 (论文配套 终极版)

核心功能：
1. 提取 32 维全局统计特征矩阵（含最大跳跃比，严格保留 6 位小数）。
2. 内嵌第五章专家评分模型：基于分段线性函数自动打分并输出一次/二次短路结论。
3. 全自动 Excel 学术级排版：大类独占一行、合并单元格、加粗灰底、小类缩进。
"""

import os
import time
import numpy as np
import pyvista as pv
import pandas as pd
from sklearn.decomposition import PCA
import warnings

# 引入 Excel 排版模块
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

warnings.filterwarnings('ignore')


def calculate_gini_coefficient(array):
    if len(array) == 0:
        return 0.0
    array = np.sort(np.array(array, dtype=np.float64))
    index = np.arange(1, array.shape[0] + 1)
    n = array.shape[0]
    return ((np.sum((2 * index - n - 1) * array)) / (n * np.sum(array)))


def format_six_decimals(x):
    if isinstance(x, (int, np.integer)):
        return x
    elif isinstance(x, (float, np.floating)):
        return round(float(x), 6)
    return x


def calculate_score(value, min_val, max_val, max_score):
    """第五章：计算分段线性得分"""
    if value < min_val:
        return 0.0
    elif value > max_val:
        return float(max_score)
    else:
        return max_score * (value - min_val) / (max_val - min_val)


def analyze_hole_parameters(vtk_file: str, output_folder: str, min_points: int = 10):
    print("=" * 65)
    print("启动导线熔痕孔洞三维特征量化与智能鉴定引擎")
    print("=" * 65)
    start_time = time.time()

    os.makedirs(output_folder, exist_ok=True)

    print(f"正在加载 VTK 模型: {vtk_file}")
    mesh = pv.read(vtk_file)
    if isinstance(mesh, pv.UnstructuredGrid):
        mesh = mesh.extract_surface()

    if "RegionId" not in mesh.cell_data:
        mesh = mesh.connectivity(largest=False)

    region_ids = mesh.cell_data["RegionId"]
    unique_regions = np.unique(region_ids)

    global_center = np.array(mesh.center)
    gb = mesh.bounds
    global_bbox_volume = (gb[1] - gb[0]) * (gb[3] - gb[2]) * (gb[5] - gb[4])

    holes_data = []

    print("正在并行解析单体孔洞的多维形态学特征...")
    for rid in unique_regions:
        mask = (region_ids == rid)
        sub_mesh = mesh.extract_cells(mask).extract_surface()

        if sub_mesh.n_points < min_points:
            continue

        points = sub_mesh.points
        V = abs(sub_mesh.volume)
        A = sub_mesh.area

        if V <= 1e-8 or A <= 1e-8:
            continue

        equivalent_diameter = (6 * V / np.pi) ** (1 / 3)
        sphericity = (np.pi ** (1 / 3) * (6 * V) ** (2 / 3)) / A

        bbox = sub_mesh.bounds
        bbox_volume = (bbox[1] - bbox[0]) * (bbox[3] - bbox[2]) * (bbox[5] - bbox[4])
        extent = V / bbox_volume if bbox_volume > 0 else 0.0

        if len(points) >= 4:
            pca = PCA(n_components=3)
            pca.fit(points)
            axis_lengths = 2 * np.sqrt(np.sort(pca.explained_variance_)[::-1])
            aspect_ratio = axis_lengths[0] / axis_lengths[2] if axis_lengths[2] > 0 else 1.0
            major_axis_len = axis_lengths[0]
        else:
            aspect_ratio = 1.0
            major_axis_len = equivalent_diameter

        local_center = np.array(sub_mesh.center)
        dist_to_center = np.linalg.norm(local_center - global_center)

        holes_data.append({
            "孔洞编号": int(rid),
            "体积 (mm³)": V,
            "表面积 (mm²)": A,
            "等效直径 (mm)": equivalent_diameter,
            "球度": sphericity,
            "矩度": extent,
            "主轴比": aspect_ratio,
            "长边 (mm)": major_axis_len,
            "中心距 (mm)": dist_to_center
        })

    df_holes = pd.DataFrame(holes_data)
    df_holes = df_holes.sort_values(by="体积 (mm³)", ascending=False).reset_index(drop=True)
    df_holes = df_holes.round(6)

    print("正在聚合全局空间分布与极化特征矩阵...")

    total_count = len(df_holes)
    if total_count == 0:
        print("未检测到有效孔洞。")
        return

    density = total_count / global_bbox_volume if global_bbox_volume > 0 else 0
    total_volume = df_holes["体积 (mm³)"].sum()
    total_area = df_holes["表面积 (mm²)"].sum()

    v_vols = df_holes["体积 (mm³)"].values
    v_areas = df_holes["表面积 (mm²)"].values
    v_diams = df_holes["等效直径 (mm)"].values
    v_spheres = df_holes["球度"].values
    v_extents = df_holes["矩度"].values
    v_aspects = df_holes["主轴比"].values
    v_majors = df_holes["长边 (mm)"].values
    v_dists = df_holes["中心距 (mm)"].values

    # ==========================================
    # 核心算法：计算最大体积跳跃比
    # ==========================================
    jump_ratios = []
    for i in range(len(v_vols) - 1):
        if v_vols[i + 1] > 0:
            jump_ratios.append(v_vols[i] / v_vols[i + 1])
    max_jump_ratio = np.max(jump_ratios) if jump_ratios else 1.0

    # 提取三大判别核心参数
    max_vol_ratio = (np.max(v_vols) / total_volume) * 100
    gini_coeff = calculate_gini_coefficient(v_vols)
    cv_coeff = np.std(v_vols) / np.mean(v_vols)

    # ==========================================
    # 第五章：专家系统多特征加权打分模型
    # ==========================================
    s_gini = calculate_score(gini_coeff, 0.4, 0.7, 40)
    s_ratio = calculate_score(max_vol_ratio, 15.0, 50.0, 30)
    s_jump = calculate_score(max_jump_ratio, 2.0, 4.0, 30)

    total_score = s_gini + s_ratio + s_jump
    prediction_result = "二次短路" if total_score >= 50 else "一次短路"

    # 原始的带有类别的数据结构 (增加专家打分模块)
    raw_global_stats = [
        {"类": "基础宏观量", "名": "孔洞总数量", "值": total_count, "单": "个"},
        {"类": "基础宏观量", "名": "孔洞空间密度", "值": density, "单": "个/mm³"},
        {"类": "基础宏观量", "名": "孔洞总体积", "值": total_volume, "单": "mm³"},
        {"类": "基础宏观量", "名": "孔洞总表面积", "值": total_area, "单": "mm²"},

        {"类": "体量特征矩阵", "名": "最大体积", "值": np.max(v_vols), "单": "mm³"},
        {"类": "体量特征矩阵", "名": "平均体积", "值": np.mean(v_vols), "单": "mm³"},
        {"类": "体量特征矩阵", "名": "最小体积", "值": np.min(v_vols), "单": "mm³"},
        {"类": "体量特征矩阵", "名": "最大表面积", "值": np.max(v_areas), "单": "mm²"},
        {"类": "体量特征矩阵", "名": "平均表面积", "值": np.mean(v_areas), "单": "mm²"},
        {"类": "体量特征矩阵", "名": "最小表面积", "值": np.min(v_areas), "单": "mm²"},
        {"类": "体量特征矩阵", "名": "最大等效直径", "值": np.max(v_diams), "单": "mm"},
        {"类": "体量特征矩阵", "名": "平均等效直径", "值": np.mean(v_diams), "单": "mm"},
        {"类": "体量特征矩阵", "名": "最小等效直径", "值": np.min(v_diams), "单": "mm"},

        {"类": "形态特征矩阵", "名": "最大球度", "值": np.max(v_spheres), "单": "-"},
        {"类": "形态特征矩阵", "名": "平均球度", "值": np.mean(v_spheres), "单": "-"},
        {"类": "形态特征矩阵", "名": "最小球度", "值": np.min(v_spheres), "单": "-"},
        {"类": "形态特征矩阵", "名": "最大矩度", "值": np.max(v_extents), "单": "-"},
        {"类": "形态特征矩阵", "名": "平均矩度", "值": np.mean(v_extents), "单": "-"},
        {"类": "形态特征矩阵", "名": "最小矩度", "值": np.min(v_extents), "单": "-"},

        {"类": "延伸与空间矩阵", "名": "最大轴比", "值": np.max(v_aspects), "单": "-"},
        {"类": "延伸与空间矩阵", "名": "平均轴比", "值": np.mean(v_aspects), "单": "-"},
        {"类": "延伸与空间矩阵", "名": "最小轴比", "值": np.min(v_aspects), "单": "-"},
        {"类": "延伸与空间矩阵", "名": "最大长边", "值": np.max(v_majors), "单": "mm"},
        {"类": "延伸与空间矩阵", "名": "平均长边", "值": np.mean(v_majors), "单": "mm"},
        {"类": "延伸与空间矩阵", "名": "最小长边", "值": np.min(v_majors), "单": "mm"},
        {"类": "延伸与空间矩阵", "名": "最大中心距", "值": np.max(v_dists), "单": "mm"},
        {"类": "延伸与空间矩阵", "名": "平均中心距", "值": np.mean(v_dists), "单": "mm"},
        {"类": "延伸与空间矩阵", "名": "最小中心距", "值": np.min(v_dists), "单": "mm"},

        {"类": "分布极化特征", "名": "最大孔洞体积占比", "值": max_vol_ratio, "单": "%"},
        {"类": "分布极化特征", "名": "体积变异系数 (CV)", "值": cv_coeff, "单": "-"},
        {"类": "分布极化特征", "名": "体积基尼系数 (Gini)", "值": gini_coeff, "单": "-"},
        {"类": "分布极化特征", "名": "最大体积跳跃比", "值": max_jump_ratio, "单": "-"},

        # 新增第五章鉴定模块输出
        {"类": "专家系统综合鉴定", "名": "基尼系数得分 (40分满)", "值": s_gini, "单": "分"},
        {"类": "专家系统综合鉴定", "名": "最大孔洞体积占比得分 (30分满)", "值": s_ratio, "单": "分"},
        {"类": "专家系统综合鉴定", "名": "跳跃比得分 (30分满)", "值": s_jump, "单": "分"},
        {"类": "专家系统综合鉴定", "名": "最终综合总分", "值": total_score, "单": "分"},
        {"类": "专家系统综合鉴定", "名": "系统预测结论", "值": prediction_result, "单": "-"}
    ]

    # ==========================================
    # 核心转换：将平面结构转换为带大类标题的层级结构
    # ==========================================
    elegant_stats = []
    current_category = ""
    for stat in raw_global_stats:
        if stat["类"] != current_category:
            current_category = stat["类"]
            elegant_stats.append({
                "参数名称": f"■ {current_category}",
                "数值": "",
                "单位": ""
            })

        elegant_stats.append({
            "参数名称": f"  {stat['名']}",
            "数值": format_six_decimals(stat['值']) if isinstance(stat['值'], (int, float)) else stat['值'],
            "单位": stat['单']
        })

    df_global_elegant = pd.DataFrame(elegant_stats)

    excel_path = os.path.join(output_folder, "智能鉴定与三维全参数矩阵.xlsx")
    csv_holes_path = os.path.join(output_folder, "单体孔洞参数明细.csv")

    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_global_elegant.to_excel(writer, sheet_name="智能鉴定与全局特征", index=False)
            df_holes.to_excel(writer, sheet_name="单体孔洞参数明细", index=False)

        wb = load_workbook(excel_path)
        ws = wb["智能鉴定与全局特征"]

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 10

        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        highlight_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # 为鉴定结果加亮

        for row in range(2, ws.max_row + 1):
            cell_A = ws[f'A{row}']
            cell_B = ws[f'B{row}']
            cell_C = ws[f'C{row}']

            if str(cell_A.value).startswith('■'):
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                cell_A.font = Font(bold=True, size=11, color="333333")

                # 为专家鉴定类目标题换个醒目颜色
                if "专家系统" in str(cell_A.value):
                    cell_A.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                else:
                    cell_A.fill = gray_fill
                cell_A.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell_B.alignment = Alignment(horizontal='center')
                cell_C.alignment = Alignment(horizontal='center')

                # 为最终结论加粗高亮
                if "系统预测结论" in str(cell_A.value) or "最终综合总分" in str(cell_A.value):
                    cell_B.font = Font(bold=True, color="C00000")
                    cell_A.fill = highlight_fill
                    cell_B.fill = highlight_fill
                    cell_C.fill = highlight_fill

        wb.save(excel_path)
        print(f"学术级排版完成！鉴定报告与特征矩阵已导出为: {excel_path}")
    except Exception as e:
        print(f"提示: Excel美化失败 ({str(e)})，可能是未安装 openpyxl 库。")

    df_holes.to_csv(csv_holes_path, index=False, encoding='utf-8-sig', float_format='%.6f')

    elapsed_time = time.time() - start_time
    print("=" * 65)
    print(f"解析与鉴定完毕！总耗时: {elapsed_time:.2f} 秒")
    print("核心判据指标提取")
    print(f"  最大孔洞体积占比 : {max_vol_ratio:8.4f} %")
    print(f"  体积基尼系数(Gini): {gini_coeff:8.4f}")
    print(f"  最大体积跳跃比   : {max_jump_ratio:8.4f}")
    print("专家系统智能打分结果:")
    print(f"  Gini 得分        : {s_gini:8.2f} / 40.0")
    print(f"  占比得分         : {s_ratio:8.2f} / 30.0")
    print(f"  跳跃比得分       : {s_jump:8.2f} / 30.0")
    print("-" * 30)
    print(f"  最终综合总分    : {total_score:8.2f} 分")
    print(f"  系统预测结论    : >>> {prediction_result} <<<")
    print("=" * 65)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='导线熔痕孔洞特征提取与智能鉴定')
    parser.add_argument('vtk_file', help='VTK 文件路径')
    parser.add_argument('--output-dir', required=True, help='输出目录路径')
    args = parser.parse_args()

    vtk_input = args.vtk_file
    output_dir = args.output_dir

    if os.path.exists(vtk_input):
        analyze_hole_parameters(vtk_input, output_dir)
    else:
        print(f"找不到文件: '{vtk_input}'")